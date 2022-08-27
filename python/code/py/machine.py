# -*- coding=utf-8 -*-
import os,sys
import openpyxl
import datetime
import datetime as dt
from datetime import date

model = './缤纷生活 - 测试组设备管理-模版.xlsx'

while True:
    Date = input('请输入需要生成文档的截止日期(不输直接回车默认当前日期）,形如 YYYYMMDD ：')
    if Date == '':#如果不输入就默认取当天的日期
        filedate = str(dt.date.today()).replace('-','')
        break
    else:
        # 判断输入是否有非法字符或者长度是否为8个数字
        if Date.isdigit() == False or len(Date) == 8:
            # 比较日期格式是否正确
            try:
                formdate = datetime.datetime.strptime(Date, '%Y%m%d')
                filedate = Date
                break
            except ValueError:
                print("输入日期的格式不合法哦，请重新检查")
        else:
            print("输入格式不合法！请按照样例格式输入日期！")

filename=model.replace('模版',str(filedate))
print('已生成文件名：',filename)

cha=datetime.datetime.strptime('20220819', '%Y%m%d')-datetime.datetime.strptime('20220815', '%Y%m%d')

# 1: 载入文件
workbook = openpyxl.load_workbook(model)
worksheet = workbook.worksheets[1]
print('读取',str(worksheet),'完成')
# 2： 第一列前插入数据
for index, row in enumerate(worksheet.rows):
    if index in [0,1,2,3]:
        print('已跳过第',str(index+1),'行...')
        pass
    else:
        print('正在写入第',str(index+1),'行...')
        worksheet.cell(index, 7, datetime.datetime.strptime(filedate, '%Y%m%d')-cha)
        worksheet.cell(index, 8, datetime.datetime.strptime(filedate, '%Y%m%d'))

# 枚举出来是tuple类型，从 0 开始计数
workbook.save(filename)

print("写入完毕")
